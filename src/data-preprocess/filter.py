# Filters the data to only include Singapore H-1B1 LCA
from openpyxl import load_workbook
from openpyxl import Workbook
import os

def filterLCAFileForH1B1Only(filename, rootPath, outputPath):
  print("Filtering " + filename + " for H-1B1 Singapore")
  # Load the workbook, appending rootPath to the filename
  wb = load_workbook(filename=rootPath + filename, read_only=True)
  ws = wb.worksheets[0]

  def isH1B1SGRow(row):
    # get the 6th cell
    h1b1type = row[5].value
    if h1b1type == "H-1B1 Singapore":
        return True
    else:
        return False

  outWB = Workbook()
  outWS = outWB.active

  write_row = 1
  row_index = 0

  for row in ws.rows:
    if isH1B1SGRow(row) or row_index == 0:
      for cell in row:
        # check if cell is EmptyCell
        if cell.value == None:
          # if so, skip
          continue
        cell_column = cell.column
        new_cell = outWS.cell(column=cell_column, row=write_row, value=cell.value)
      write_row += 1
    row_index += 1

  # Print length of filtered data
  print("Filtered data has " + str(write_row - 1) + " rows")

  # strip .xlsx from the filename
  filename = filename[:-5]
  # append -h1b1-only.xlsx to the filename
  filename = filename + "-h1b1-only.xlsx"
  # Save the workbook
  outWB.save(outputPath + filename)

ROOT_PATH = "data/"
OUTPUT_PATH = "filtered-data/"

# Iterate through each .xlsx file in the root path, and filter it
for filename in os.listdir(ROOT_PATH):
  if filename.endswith(".xlsx"):
    filterLCAFileForH1B1Only(filename, ROOT_PATH, OUTPUT_PATH)
    continue
  else:
    continue
